import openpyxl
import os

def check_file(path):
    if not os.path.exists(path):
        print(f"File not found: {path}")
        return
    print(f"\n--- Checking file: {path} ---")
    wb = openpyxl.load_workbook(path, data_only=True)
    print("Sheets:", wb.sheetnames)
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"Sheet '{name}' dimensions: {ws.dimensions}")
        # Search for 'effectif' in the first 100 rows and 20 columns
        for r in range(1, min(ws.max_row + 1, 150)):
            for c in range(1, min(ws.max_column + 1, 30)):
                val = ws.cell(row=r, column=c).value
                if val and any(x in str(val).lower() for x in ['effectif', 'nb enfant', 'nombre enfant', 'total enfant', 'nbre']):
                    print(f"[{name}] Found keyword at Row {r}, Col {c} ({openpyxl.utils.get_column_letter(c)}): {val}")
                    # Print neighboring cells
                    row_vals = [ws.cell(row=r, column=i).value for i in range(max(1, c-2), min(ws.max_column+1, c+10))]
                    print(f"  Row {r} around: {row_vals}")

check_file(r"C:\Users\stagedg2\Projet_mairie_outil_tarification\Donnees\Autres\CALC DEP(4).xlsx")
check_file(r"C:\Users\stagedg2\Projet_mairie_outil_tarification\tarification-api\VF_REC_DEP.xlsx")
check_file(r"C:\Users\stagedg2\Projet_mairie_outil_tarification\VF_REC_DEP.xlsx")
check_file(r"C:\Users\stagedg2\Projet_mairie_outil_tarification\Donnees\Autres\Classeur pour la tarification.xlsx")
