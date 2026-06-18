import openpyxl

wb = openpyxl.load_workbook(r"c:\Users\stagedg2\Projet_mairie_outil_tarification\VF_REC_DEP.xlsx", data_only=True)
print("Sheets:", wb.sheetnames)

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    print(f"\n--- Sheet: {sheetname} ({ws.max_row} rows, {ws.max_column} cols) ---")
    # Print first 15 rows, 10 columns
    for r in range(1, min(ws.max_row + 1, 30)):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column + 1, 15))]
        # print if not all None
        if any(v is not None for v in row_vals):
            print(f"Row {r:02d}: {row_vals}")
