import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\stagedg2\Projet_mairie_outil_tarification\Donnees\Autres\CALC DEP(4).xlsx", data_only=True)
ws = wb["syntheses charges"]

print("syntheses charges sheet Rows 25 to 45:")
for r in range(25, 46):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, 12)]
    print(f"Row {r:02d}: {row_vals}")

print("\nDEP sheet of VF_REC_DEP.xlsx:")
wb2 = openpyxl.load_workbook(r"C:\Users\stagedg2\Projet_mairie_outil_tarification\VF_REC_DEP.xlsx", data_only=True)
ws2 = wb2["DEP"]
for r in range(1, 70):
    # Print lines that look like total or contain values for effectifs
    row_vals = [ws2.cell(row=r, column=c).value for c in range(1, 23)]
    # Filter empty rows
    if any(row_vals):
        # print first few columns and the ones we know
        print(f"Row {r:02d}: {row_vals[:12]}")
