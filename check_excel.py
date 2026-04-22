import openpyxl
wb = openpyxl.load_workbook(r"C:\Users\stagedg2\Projet_mairie_outil_tarification\Donnees\Autres\CALC DEP(4).xlsx", data_only=True)
ws = wb["syntheses charges"]

print("Row 23 values:")
for col in range(1, 10):
    print(f"Col {col} : {ws.cell(row=23, column=col).value}")

print("\nSpecific Séjours cell:")
# Séjours is column 9 (I)
print(ws.cell(row=23, column=9).value)
