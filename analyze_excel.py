import openpyxl

# Analyser le fichier CALC DEP (3).xlsx
wb = openpyxl.load_workbook('Donnees/Autres/CALC DEP (3).xlsx')

print('=== FEUILLES DU CLASSEUR ===')
for i, sheet_name in enumerate(wb.sheetnames):
    print(f'{i}: {sheet_name}')

print('\n=== ANALYSE DETAILLEE ===')

# Analyser chaque feuille
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f'\n--- FEUILLE: {sheet_name} ---')
    print(f'Dimensions: {sheet.dimensions}')
    print(f'Lignes: {sheet.max_row}, Colonnes: {sheet.max_column}')

    # Afficher les premières lignes pour comprendre la structure
    print('Premières lignes:')
    for row in range(1, min(6, sheet.max_row + 1)):
        row_data = []
        for col in range(1, min(8, sheet.max_column + 1)):
            cell = sheet.cell(row=row, column=col)
            value = cell.value
            if value is not None:
                # Tronquer les longues chaînes
                str_value = str(value)
                if len(str_value) > 30:
                    str_value = str_value[:27] + '...'
                row_data.append(str_value)
            else:
                row_data.append('')
        print(f'Ligne {row}: {row_data}')

print('\n=== ANALYSE TERMINEE ===')