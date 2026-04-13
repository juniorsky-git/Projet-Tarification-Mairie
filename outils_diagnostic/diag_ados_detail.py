import openpyxl
import os

FICHIER = "Donnees/Autres/CALC DEP.xlsx"

def diagnostiquer_ados():
    print("=================================================")
    print("   DIAGNOSTIC FINAL : DEPENSES ESPACE ADOS      ")
    print("=================================================\n")

    if not os.path.exists(FICHIER):
        print(f"Erreur : {FICHIER} introuvable.")
        return

    try:
        # data_only=True permet de voir les resultats des calculs (les chiffres)
        wb = openpyxl.load_workbook(FICHIER, data_only=True, read_only=True)
        ws = wb["Simulation"]

        # Ligne 74 : Les noms des colonnes (index 1-based dans Excel donc 74)
        # Ligne 75 : Les valeurs reelles
        headers = [cell.value for cell in ws[74]]
        values = [cell.value for cell in ws[75]]

        print(f"{'NATURE DE LA DEPENSE':<40} | {'MONTANT':<15}")
        print("-" * 60)

        # On parcourt toutes les colonnes à partir de la colonne C (index 2)
        total_verifie = 0
        for i in range(2, len(headers)):
            label = str(headers[i]).replace("\n", " ").strip() if headers[i] else "???"
            val = values[i] if i < len(values) else 0
            
            if val and val != 0:
                # Si c'est le total general, on le met en evidence
                if "TOTAL" in label.upper():
                    print("-" * 60)
                    print(f"{label:<40} | {val:>12.2f} €")
                else:
                    print(f"{label:<40} | {val:>12.2f} €")
                    # On additionne pour verifier
                    if isinstance(val, (int, float)):
                        total_verifie += val

        print("\n" + "=" * 60)
        print(f"VERIFICATION CALCUL : {total_verifie:.2f} €")
        print("=================================================")

    except Exception as e:
        print(f"Erreur technique : {e}")

if __name__ == "__main__":
    diagnostiquer_ados()
