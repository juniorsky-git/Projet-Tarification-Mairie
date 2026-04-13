import openpyxl
import os

# Chemin vers le fichier Excel
FICHIER = "Donnees/Autres/CALC DEP.xlsx"

def analyser_conso_eau():
    print("=== DIAGNOSTIC PYTHON : CONSO EAU ===")
    
    if not os.path.exists(FICHIER):
        print(f"Erreur : Le fichier {FICHIER} est introuvable.")
        return

    try:
        # Chargement du classeur (data_only=True pour recuperer les resultats des formules)
        print(f"Chargement de {FICHIER}...")
        wb = openpyxl.load_workbook(FICHIER, data_only=True, read_only=True)
        
        # Acces a l'onglet
        nom_onglet = "Conso eau"
        if nom_onglet not in wb.sheetnames:
            print(f"Erreur : L'onglet '{nom_onglet}' n'existe pas.")
            print("Onglets disponibles :", wb.sheetnames)
            return
            
        ws = wb[nom_onglet]
        
        # Parcours des lignes
        print(f"Analyse de l'onglet '{nom_onglet}' :")
        for i, row in enumerate(ws.iter_rows(), start=1):
            # En Python/openpyxl, row est un tuple de cellules
            nb_colonnes = len(row)
            
            # Recherche de la derniere cellule non vide pour voir la "vraie" fin
            derniere_valeur = "Vide"
            for cell in reversed(row):
                if cell.value is not None:
                    derniere_valeur = cell.value
                    break
            
            print(f"Ligne {i:02d} : {nb_colonnes} colonnes | Derniere valeur utile : {derniere_valeur}")
            
            # On s'arrete apres 10 lignes pour la demonstration
            if i >= 10:
                print("... (suite du fichier ignoree pour l'exemple) ...")
                break

    except Exception as e:
        print(f"Une erreur est survenue : {e}")

if __name__ == "__main__":
    analyser_conso_eau()
