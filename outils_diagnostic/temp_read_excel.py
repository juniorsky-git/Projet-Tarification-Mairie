from pathlib import Path
import openpyxl

p = Path('Donnees/Autres/CALC DEP.xlsx')
print('exists', p.exists())
wb = openpyxl.load_workbook(p, data_only=True)
sh = wb.worksheets[0]
print('sheet0', sh.title)
print('R47 cols A-R', [sh.cell(row=47, column=j).value for j in range(1, 19)])

with open('temp_excel_out.txt', 'w', encoding='utf-8') as f:
    f.write('sheet0 ' + str(sh.title) + '\n')
    f.write('R47:' + str([sh.cell(row=47, column=j).value for j in range(1, 19)]) + '\n')
    f.write('max_row:' + str(sh.max_row) + '\n')
    for i in range(1, sh.max_row + 1):
        row = [sh.cell(row=i, column=j).value for j in range(1, 21)]
        lib = str(row[3]).upper() if row[3] is not None else ''
        ser = str(row[18]).upper() if row[18] is not None else ''
        ant = str(row[19]).upper() if row[19] is not None else ''
        if 'LOISIR' in lib or 'LOISIRS' in lib or 'ACCUEIL' in lib or 'RESTGAV' in ant or '2-RE' in ser:
            f.write('ROW ' + str(i) + ' ANT ' + ant + ' SER ' + ser + ' LIB ' + str(row[3]) + ' H ' + str(row[7]) + ' R ' + str(row[17]) + '\n')
