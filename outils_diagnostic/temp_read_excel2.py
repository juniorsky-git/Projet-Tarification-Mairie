from pathlib import Path
import openpyxl

p = Path('Donnees/Autres/CALC DEP.xlsx')
wb = openpyxl.load_workbook(p, data_only=True)
sh = wb.worksheets[0]

print('sheet0', sh.title)
print('R47 D:', sh.cell(row=47, column=4).value)
print('R47 R:', sh.cell(row=47, column=18).value)
print('R47 S:', sh.cell(row=47, column=19).value)
print('R47 T:', sh.cell(row=47, column=20).value)

for i in range(1, sh.max_row + 1):
    lib = sh.cell(row=i, column=4).value
    if lib is None:
        continue
    lib_str = str(lib).upper()
    if 'LOISIR' in lib_str or 'ACCUEIL' in lib_str:
        print('ROW', i, 'LIB', lib_str)
        print('  H', sh.cell(row=i, column=8).value, 'R', sh.cell(row=i, column=18).value, 'S', sh.cell(row=i, column=19).value, 'T', sh.cell(row=i, column=20).value)
