from pathlib import Path
import openpyxl

p = Path('Donnees/Autres/CALC DEP.xlsx')
wb = openpyxl.load_workbook(p, data_only=True)
sh = wb.worksheets[8]

with open('temp_simulation_out.txt', 'w', encoding='utf-8') as out:
    out.write('sheet ' + str(sh.title) + '\n')
    out.write('max_row ' + str(sh.max_row) + ' max_col ' + str(sh.max_column) + '\n')

    keywords = ['MDJ', 'PTIT PRINCE', 'CLGAV', 'CLJP1', 'CLLMICH', 'ACCUEIL', 'LOISIR']
    filtered = []
    for i in range(1, sh.max_row + 1):
        row = [sh.cell(row=i, column=j).value for j in range(1, sh.max_column + 1)]
        text = ' '.join([str(x).upper() for x in row if x is not None])
        if any(k in text for k in keywords):
            filtered.append((i, row, text))

    out.write('found ' + str(len(filtered)) + ' rows\n')
    for i, row, text in filtered:
        out.write('ROW ' + str(i) + ' ' + ' | '.join([str(x) for x in row]) + '\n')
print('done')
