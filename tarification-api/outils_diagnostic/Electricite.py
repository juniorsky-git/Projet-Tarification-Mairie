import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generer_synthese_elec():
    # 1. Chargement des données brutes
    # On lit le fichier sans en-tête car la structure est irrégulière
    df_raw = pd.read_csv('CALC DEP(4).xlsx - CONSO ELEC.csv', header=None)

    data_rows = []

    # 2. Logique d'extraction
    # Les blocs de facturation commencent à la colonne index 7
    # Chaque bloc contient 4 colonnes : [Période Conso, Période Abo, Consos kWh, Montant TTC]
    for col_base in range(7, df_raw.shape[1] - 3, 4):
        # On récupère les infos de facture situées en haut (Ligne 2 et 3)
        date_f = df_raw.iloc[1, col_base]
        num_f = df_raw.iloc[2, col_base]
        
        # On saute le bloc si les infos de facture sont vides
        if pd.isna(date_f) and pd.isna(num_f):
            continue
            
        # On parcourt les lignes de sites (à partir de la ligne 10 / index 9)
        for row_idx in range(9, df_raw.shape[0]):
            site_nom = df_raw.iloc[row_idx, 0]   # Nom du site
            pdl_ref = df_raw.iloc[row_idx, 1]    # Référence RAE/PCE
            adresse = df_raw.iloc[row_idx, 5]    # Adresse Ville
            
            if pd.isna(pdl_ref) and pd.isna(adresse):
                continue
                
            # Extraction des données du bloc actuel
            p_conso = df_raw.iloc[row_idx, col_base]
            p_abo = df_raw.iloc[row_idx, col_base + 1]
            conso_kwh = df_raw.iloc[row_idx, col_base + 2]
            montant = df_raw.iloc[row_idx, col_base + 3]
            
            # On n'ajoute la ligne que s'il y a une consommation réelle (différente de 0)
            val_conso = str(conso_kwh).replace(',', '.').strip()
            if val_conso not in ["", "0", "0.0", "nan", "NaN"]:
                data_rows.append({
                    "NOM - SITE": site_nom,
                    "REFERENCE PDL": pdl_ref,
                    "ADRESSE (CROSNE)": adresse,
                    "DATE FACTURE": date_f,
                    "N° DE FACTURE": num_f,
                    "Période de consos": p_conso,
                    "Période d'abonnement": p_abo,
                    "Consos en kWh": pd.to_numeric(val_conso, errors='coerce'),
                    "Montant total en € TTC": pd.to_numeric(str(montant).replace(',', '.').strip(), errors='coerce')
                })

    # 3. Création du DataFrame et export Excel
    df_final = pd.DataFrame(data_rows)
    file_name = "Synthese_ELEC_Automatique.xlsx"
    writer = pd.ExcelWriter(file_name, engine='openpyxl')
    df_final.to_excel(writer, sheet_name='ELEC Détail', index=False)

    # 4. Mise en forme (Couleurs et Règle du "ET")
    ws = writer.sheets['ELEC Détail']
    
    blue_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    red_alert = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    white_font = Font(bold=True, color="FFFFFF")
    red_font = Font(color="9C0006", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Style des titres
    for cell in ws[1]:
        cell.fill = blue_header
        cell.font = white_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Style des données + Règle "et"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # On vérifie si " et " est dans le Nom (col 1) ou l'Adresse (col 3)
        nom_val = str(row[0].value).lower()
        addr_val = str(row[2].value).lower()
        
        if " et " in f" {nom_val} " or " et " in f" {addr_val} ":
            row[0].fill = red_alert
            row[0].font = red_font
            row[2].fill = red_alert
            row[2].font = red_font
            
        for cell in row:
            cell.border = thin_border
            # Formatage des colonnes Consos et Montant
            if cell.column in [8, 9]:
                cell.number_format = '#,##0.00'

    # Ajustement largeur colonnes
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    writer.close()
    print(f"Fichier généré : {file_name}")

# Lancement
generer_synthese_elec()