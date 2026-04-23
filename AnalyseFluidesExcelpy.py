import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- CONFIGURATION DES STYLES ---
def apply_pro_style(ws, header_color, col_to_check_et=None):
    """Applique un style professionnel et la règle du 'et' en rouge."""
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    red_font = Font(color="9C0006", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Style des en-têtes (Ligne 1)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Traitement des lignes de données
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # Règle du "et" pour la colonne d'adresse spécifiée
        if col_to_check_et:
            cell_addr = row[col_to_check_et - 1] # -1 car index 0
            val = str(cell_addr.value).lower()
            if " et " in f" {val} ":
                cell_addr.fill = red_fill
                cell_addr.font = red_font

        # Bordures et formats numériques pour toute la ligne
        for cell in row:
            cell.border = thin_border
            # Format monétaire/nombre pour les colonnes de conso et montant
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

    # Ajustement automatique de la largeur des colonnes
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

# --- 1. TRAITEMENT DE L'EAU ---
def process_water():
    df = pd.read_csv('CALC DEP(4).xlsx - Conso eau.csv', header=None)
    # On sélectionne les colonnes importantes
    cols = [0, 2, 3, 5, 6, 7, 8, 9, 10, 15, 16, 17, 18, 19, 20]
    df_clean = df.iloc[5:, cols].copy()
    df_clean.columns = ['PDL', 'TYPE', 'ADRESSE', 'S1_PERIODE', 'S1_ABO', 'S1_M3', 'S1_EUR', 'S1_DATE', 'S1_FACT',
                        'S2_PERIODE', 'S2_ABO', 'S2_M3', 'S2_EUR', 'S2_DATE', 'S2_FACT']
    
    writer = pd.ExcelWriter('Synthese_Eau_Detaillee.xlsx', engine='openpyxl')
    df_clean.to_excel(writer, index=False, sheet_name='Eau')
    apply_pro_style(writer.sheets['Eau'], "1F4E78", col_to_check_et=3) # Bleu
    writer.close()

# --- 2. TRAITEMENT DU GAZ ---
def process_gas():
    df_raw = pd.read_csv('CALC DEP(4).xlsx - CONSO GAZ.csv', header=None)
    data_rows = []

    # Le fichier Gaz est horizontal, on boucle sur les blocs de factures (tous les 4 colonnes)
    for col_start in range(4, df_raw.shape[1] - 3, 4):
        date_f = df_raw.iloc[1, col_start]
        num_f = df_raw.iloc[2, col_start]
        fourn = df_raw.iloc[3, col_start]
        
        if pd.isna(date_f) and pd.isna(num_f): continue
            
        for row_idx in range(9, df_raw.shape[0]):
            addr_crosne = df_raw.iloc[row_idx, 2]
            if pd.isna(addr_crosne): continue
            
            conso = pd.to_numeric(str(df_raw.iloc[row_idx, col_start+2]).replace(',','.'), errors='coerce')
            montant = pd.to_numeric(str(df_raw.iloc[row_idx, col_start+3]).replace(',','.'), errors='coerce')
            
            if pd.notna(conso) or pd.notna(montant):
                data_rows.append({
                    "ADRESSE": addr_crosne,
                    "DATE FACTURE": date_f,
                    "N° FACTURE": num_f,
                    "FOURNISSEUR": fourn,
                    "CONSO M3": conso,
                    "MONTANT TTC": montant
                })

    df_final = pd.DataFrame(data_rows)
    writer = pd.ExcelWriter('Synthese_Gaz_Detaillee.xlsx', engine='openpyxl')
    df_final.to_excel(writer, index=False, sheet_name='Gaz')
    apply_pro_style(writer.sheets['Gaz'], "ED7D31", col_to_check_et=1) # Orange
    writer.close()

# Exécution
process_water()
process_gas()
print("Fichiers générés avec succès !")